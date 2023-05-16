from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
# from django.db.models.functions import Outer
from django.contrib.auth.models import UserManager

from django.views.decorators.csrf import csrf_exempt

from .models import CBCodeHdr, CBCodeDtl, CMItem, CBItemgrp, CMItemaccnt, CMFactory, CMBOM

import openpyxl
from openpyxl import Workbook

import os
from django.core.files.storage import FileSystemStorage

from datetime import datetime

import requests
import json

import pymysql

from django.conf import settings
MYDB = getattr(settings, "DATABASES", None)
MYDB_NAME = MYDB["default"]["NAME"]
MYDB_USER = MYDB["default"]["USER"]
MYDB_PWD = MYDB["default"]["PASSWORD"]
MYDB_HOST = MYDB["default"]["HOST"]
dbCon = pymysql.connect(host=MYDB_HOST, user=MYDB_USER, password=MYDB_PWD, database=MYDB_NAME)

# dbCon = pymysql.connect('223.194.46.212','eduuser','12345!','edudb')

def home(request):
    context = {}

    # request.session['member_name'] = "홍길동"
    
    context["title"] = "Main"
    context["result_msg"] = "Django Application Framework"

    return render(request, "home.html", context)


# *********************************************************************************************************************
# 통합코드 관리
# *********************************************************************************************************************

def codemanage(request):
    context = {}

    if 'type_cd' in request.GET:
        typecd = request.GET['type_cd']
        rsCode = CBCodeDtl.objects.filter(type_cd=typecd, usage_flag='1')
    else:
        typecd = None
        rsCode = None

    # print(typecd)
    context["typecd"] = typecd

    rsHeader = CBCodeHdr.objects.filter(usage_flag='1')

    context["rsHeader"] = rsHeader
    context["rsCode"] = rsCode

    print(rsCode)

    context["title"] = "통합코드 관리"
    context["result_msg"] = "통합코드 관리"

    return render(request, "codemanage.html", context)


@csrf_exempt
def codetype_insert(request):
    context = {}

    typecd = request.GET['typecd']
    typename = request.GET['typename']

    if CBCodeHdr.objects.filter(type_cd=typecd).exists():
        context["flag"] = "1"
        context["result_msg"] = "Type이 있습니다."
        return JsonResponse(context, content_type="application/json")

    else:
        CBCodeHdr.objects.create(type_cd=typecd,
                                 type_name=typename)

        context["flag"] = "0"
        context["result_msg"] = "Type등록 성공..."
        return JsonResponse(context, content_type="application/json")

@csrf_exempt
def codetype_update(request):
    context = {}

    typeid = request.GET['typeid']
    typename = request.GET['typename']

    if CBCodeHdr.objects.filter(type_name=typename).exists():
        context["flag"] = "1"
        context["result_msg"] = "Type 명칭 중복..."
        return JsonResponse(context, content_type="application/json")
    else:
        rs = CBCodeHdr.objects.get(id=typeid)
        rs.type_name = typename
        rs.save()

        context["flag"] = "0"
        context["result_msg"] = "Type 수정 성공..."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def codetype_delete(request):
    context = {}

    typeid = request.GET['typeid']

    rs = CBCodeHdr.objects.get(id=typeid)
    typecd = rs.type_cd

    if CBCodeDtl.objects.filter(type_cd=typecd).exists():
        context["flag"] = "1"
        context["result_msg"] = "Type에 하위 코드가 있습니다. 삭제 불가..."
        return JsonResponse(context, content_type="application/json")
    else:
        rs.usage_flag = '0'
        rs.save()

        context["flag"] = "0"
        context["result_msg"] = "Type 삭제 되었습니다."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def code_insert(request):
    context = {}

    typecd = request.GET['typecd']
    codecd = request.GET['codecd']
    codename = request.GET['codename']

    if CBCodeDtl.objects.filter(type_cd=typecd, code_cd=codecd).exists():
        context["flag"] = "1"
        context["result_msg"] = "Code 중복..."
        return JsonResponse(context, content_type="application/json")
    else:
        CBCodeDtl.objects.create(type_cd=typecd,
                                 code_cd=codecd,
                                 code_name=codename)

        context["flag"] = "0"
        context["result_msg"] = "Code 등록 성공..."
        return JsonResponse(context, content_type="application/json")

@csrf_exempt
def code_update(request):
    context = {}

    codeid = request.GET['codeid']
    cd_nm = request.GET['cd_nm']
    print(cd_nm)

    rs = CBCodeDtl.objects.get(id=codeid)
    rs.cd_nm = cd_nm
    rs.save()

    context["result_msg"] = "Code Update success..."

    return JsonResponse(context, content_type="application/json")

def code_view(request):
    context = {}

    codeid = request.GET['codeid']
    rsCode = CBCodeDtl.objects.get(id=codeid)

    print(rsCode)

    context["type_cd"] = rsCode.type_cd
    context["code_cd"] = rsCode.code_cd
    context["code_name"] = rsCode.code_name

    context["title"] = "코드 VIEW"
    context["result_msg"] = "Code detail"
    return render(request, "codeview.html", context)

@csrf_exempt
def code_delete(request):
    context = {}

    codeid = request.GET['codeid']

    if CBCodeDtl.objects.get(id=codeid):
        rs = CBCodeDtl.objects.get(id=codeid)
        rs.usage_flag = '0'
        rs.save()

    context["flag"] = "0"
    context["result_msg"] = "Code deleted... "
    return JsonResponse(context, content_type="application/json")

# *********************************************************************************************************************
# 품목관리
# *********************************************************************************************************************

def itemmanage(request):
    context = {}

    strsql = "SELECT a.*, b.*, c.*, d.*, e.* " +\
             "FROM (SELECT * FROM cm_item WHERE usage_flag = '1') a " +\
             "LEFT JOIN cm_factory b ON a.factory_id = b.id " +\
             "LEFT JOIN (SELECT id, code_cd AS unit_cd, code_name AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c ON a.unit_id = c.id " +\
             "LEFT JOIN cb_itemgrp d ON a.itemgrp_id = d.id " +\
             "LEFT JOIN cm_itemaccnt e ON a.itemaccnt_id = e.id "

    rsItem = CMItem.objects.raw(strsql)[:30]
    # rsItem = CMItem.objects.filter(usage_flag='1')
    context["rsItem"] = rsItem[:100]

    rsItemgrp = CBItemgrp.objects.filter()
    rsItemaccnt = CMItemaccnt.objects.filter()
    rsFactory = CMFactory.objects.filter()
    rsUnit = CBCodeDtl.objects.filter(type_cd='unit', usage_flag='1')

    context["rsItemgrp"] = rsItemgrp
    context["rsItemaccnt"] = rsItemaccnt
    context["rsFactory"] = rsFactory
    context["rsUnit"] = rsUnit

    context["title"] = "품목관리"
    context["result_msg"] = "품목 관리"
    return render(request, "itemmanage.html", context)


@csrf_exempt
def itemtemplate_download(request):
    context = {}

    strsql1 = "SHOW TABLES LIKE 'cm_item'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW COLUMNS FROM cm_item"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

        idx = 1

        bookin = Workbook()
        sheet_in = bookin.active

        for i in rsColumns:
            sheet_in.cell(row=1, column=idx).value = i[1]
            sheet_in.cell(row=2, column=idx).value = i[0]
            idx += 1

        filename = "static/datatemplates/cm_item.xlsx"
        bookin.save(filename)
        bookin.close()

        context["template_url"] = "/static/datatemplates/cmitem.xlsx"

    else:
        context["flag"] = "1"
        context["result_msg"] = "품목코드 테이블 없음... "
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "Template downloaded... "
    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def itemdata_upload(request):
    context = {}

    print('1. Data file upload >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

    if request.method == "POST":
        uploaded_file = request.FILES['ufile']
        name_old = uploaded_file.name
        name_ext = os.path.splitext(name_old)[1]
        name_new = 'cm_item'

        file_name = name_new + name_ext
        fs = FileSystemStorage(location='static/dataupload')
        # excel의 경우 사용중 프로세스 가능성?
        if (fs.exists(file_name)):
            fs.delete(file_name)

        name = fs.save(file_name, uploaded_file)

        context['upload_url'] = fs.url(name)
        context['upload_flag'] = 'USuccess'

    else:
        return redirect("itemmanage")


    print('2. Read table columns >>>>>>>>>datasave>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

    # Table 있는지 읽어본다
    strsql1 = "SHOW TABLES LIKE 'cm_item'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTbl = cursor1.fetchone()
    cursor1.close()

    if rsTbl:
        # Table 있으면 Column을 읽는다
        strsql1 = "SHOW COLUMNS FROM cm_item"
        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

    else:
        return redirect("itemmanage")

    max_col = len(rsColumns)

    strtop = ""
    metaTmp2 = []
    cnt_col = 0

    # Meta column을 Concatenate
    # 1: column_name, 2:Type, 3:display_flag
    for i in rsColumns:
        metaTmp2.append(i[1])
        strtop += str(i[0]) + ","
        cnt_col += 1

    strtop = strtop[:-1]

    print('3. Data insert >>>>>>>>>>>>>>>datasave>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')

    filename = "static/dataupload/" + file_name

    if not os.path.isfile(filename):
        return redirect("itemmanage")

    else:
        book = openpyxl.load_workbook(filename, read_only=True)

        sheet = book.active
        max_row = sheet.max_row
        timenow = datetime.now()

        if max_row > 2:
            datacnt = 0
            for j in range(3, max_row + 1):
                lstTmp = []
                strbottom = ""
                datacnt += 1

                cnt_col = 0
                for i in range(1, max_col + 1):
                    valTmp = sheet.cell(row=j, column=i).value

                    cnt_col += 1
                    lstTmp.append(valTmp)

                    if metaTmp2[i-1][:7] == "varchar":
                        if valTmp == 'NULL':
                            strbottom += "'',"
                        elif valTmp == '':
                            strbottom += "'',"
                        elif valTmp == None:
                            strbottom += "'',"
                        else:
                            valTmpTo = str(valTmp).strip()
                            valTmpTo = valTmpTo.replace("'",'"')
                            if(len(valTmpTo) > 255):
                                valvalTmpTo2 = valTmpTo[0:255]
                            else:
                                valvalTmpTo2 = valTmpTo
                            strbottom += "'" + valvalTmpTo2 + "',"

                    elif metaTmp2[i-1][:8] == "datetime":
                        strbottom += "'" + str(timenow) + "',"

                    else:
                        if valTmp == 'NULL':
                            strbottom += "0,"
                        elif valTmp == None:
                            strbottom += "0,"
                        else:
                            valTmpTo = str(valTmp).strip()
                            valTmpTo = valTmpTo.replace("'",'"')
                            if(len(valTmpTo) > 255):
                                valvalTmpTo2 = valTmpTo[0:255]
                            else:
                                valvalTmpTo2 = valTmpTo
                            strbottom += "'" + valvalTmpTo2 + "',"

                strbottom = strbottom[:-1]

                strSql = "INSERT INTO cm_item (" + strtop + ") VALUES (" + strbottom + ")"

                print(strSql)

                if datacnt % 50 == 0:
                    print('Data insert processing --> ' + str(datacnt) + ' / ' + str(max_row - 2))

                c4 = dbCon.cursor()
                c4.execute(strSql)
                rows = c4.fetchone()
                c4.close()

            dbCon.commit()

        else:
            return redirect("itemmanage")

    return redirect("itemmanage")

def apitest(request):
    context = {}

    import pandas as pd

    query1 = """
             {
              lmetaList(ledgerNo: 1) {
                id
                ledgerNo
                columnName
                columnType
              }
            }
            """

    # API_URL = 'http://' + MYDB_HOST + ':8122/graphql'
    API_URL = 'http://223.194.46.212:8122/graphql'

    result = requests.get(API_URL, json={'query': query1})
    print(type(result))

    json_data = json.loads(result.text)
    #print(type(json_data))

    df_data = json_data['data']['lmetaList']
    print(type(df_data))

    df = pd.DataFrame(df_data)
    df = df[df.columns[::-1]]
    #print(type(df))

    rsAnalysis = [tuple(r) for r in df.to_numpy()]
    print(type(rsAnalysis))

    context["rsAnalysis"] = rsAnalysis

    context["title"] = "API test"
    context["result_msg"] = "API test"
    return render(request, "apitest.html", context)

def bommanage(request):
    context = {}

    context['itemid'] = 0
    context['itemcd'] = ""
    context['itemname'] = ""
    context['itemspec'] = ""
    context['registerdate'] = ""
    itemid = "0"
    if 'itemid' in request.GET:
        itemid = request.GET['itemid']
        context['itemid'] = itemid
        if CMItem.objects.filter(id=itemid).exists():
            rsTmp = CMItem.objects.get(id=itemid)
            context['itemcd'] = rsTmp.item_cd
            context['itemname'] = rsTmp.item_name
            context['itemspec'] = rsTmp.item_spec
            context['itemcd'] = rsTmp.item_cd
            context['registerdate'] = rsTmp.register_date

    bomid = "0"
    context['moitembase'] = 0.0
    context['jaitembase'] = 0.0
    context['unitproduct'] = '단위'
    context['lossproduct'] = 0.0
    context['demandamt'] = 0.0
    context['startdate'] = ''
    context['enddate'] = ''
    if 'bomid' in request.GET:
        bomid = request.GET['bomid']
        if CMBOM.objects.filter(id=bomid).exists():
            rsTmp2 = CMBOM.objects.get(id=bomid)
            context['moitembase'] = rsTmp2.moitem_base
            context['jaitembase'] = rsTmp2.jaitem_base
            context['unitproduct'] = rsTmp2.unit_product
            context['lossproduct'] = rsTmp2.loss_product
            context['demandamt'] = rsTmp2.demand_amt
            context['startdate'] = rsTmp2.start_date
            context['enddate'] = rsTmp2.end_date

    itemgrpid = ""
    if 'itemgrpid' in request.GET:
        itemgrpid = request.GET['itemgrpid']

    searchcode = ""
    if 'itemcode' in request.GET:
        searchcode = request.GET['itemcode']

    searchspec = ""
    if 'itemspec' in request.GET:
        searchspec = request.GET['itemspec']

    if searchcode != "":
        # rsItem = CMItem.objects.filter(Q(item_name__contains=url) | Q(url_title__contains=url)).order_by('-register_date')[:200]
        rsItem = CMItem.objects.filter(Q(item_cd__contains=searchcode))[:100]
    elif searchspec != "":
        rsItem = CMItem.objects.filter(Q(item_spec__contains=searchspec))[:100]
    elif itemgrpid != "":
        rsItem = CMItem.objects.filter(itemgrp_id=itemgrpid)[:100]
    else:
        strsql = "SELECT a.*, b.*, d.*" + \
                 "FROM (SELECT * FROM cm_item WHERE usage_flag = '1') a " + \
                 "LEFT JOIN cm_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN cb_itemgrp d ON a.itemgrp_id = d.id "
        rsItem = CMItem.objects.raw(strsql)[:100]

    context["rsItem"] = rsItem

    # strsql = f"SELECT a.*, b.* FROM cm_bom a LEFT JOIN cm_item b ON a.item_id = b.id WHERE a.top_id = {itemid}"
    # rsBOM = CMBOM.objects.raw(strsql)
    rsBOM = CMBOM.objects.filter(top_id=itemid).select_related("item")

    context["rsBOM"] = rsBOM

    rsItemgrp = CBItemgrp.objects.filter()
    context["rsItemgrp"] = rsItemgrp

    context['bomid'] = bomid
    context["itemgrpid"] = itemgrpid
    context["title"] = "BOM관리"
    context["result_msg"] = "BOM 관리"
    return render(request, "bommanage.html", context)


@csrf_exempt
def bom_create(request):
    context = {}

    itemid = request.GET['itemid']

    if CMBOM.objects.filter(item_id=itemid,parent_id=0).exists():
        print('이미 있습니다.')
        context["flag"] = "1"
        context["result_msg"] = "이미 있습니다."
        return JsonResponse(context, content_type="application/json")
    else:
        CMBOM.objects.create(bom_type='MBOM',
                             item_id=itemid,
                             parent_id=0,
                             top_id=itemid,
                             bom_order=1,
                             bom_level=0,
                             leaf_flag='0',
                             moitem_base=0.0,
                             jaitem_base=0.0,
                             unit_product='',
                             demand_amt=0.0,
                             free_flag='0',
                             loss_product=0.0,
                             start_date='',
                             end_date='',
                             register_date=datetime.now(),
                             usage_flag='1')

        rsItem = CMItem.objects.get(id=itemid)
        rsItem.bom_flag = '1'
        rsItem.save()

        context["flag"] = "0"
        context["result_msg"] = "Top level 등록 성공..."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bomitem_read(request):
    context = {}

    bomid = request.GET['bomid']
    itmtext = request.GET['itmtext']

    if itmtext == "":
        rsItem = CMItem.objects.filter(usage_flag='1')[:10]

    else:
        # strsql = f"SELECT * FROM cm_item WHERE item_cd LIKE '%{itmtext}%' or item_spec LIKE '%{itmtext}%'"
        # print(strsql)
        # rsItem = CMItem.objects.raw(strsql)
        rsItem = CMItem.objects.filter(Q(item_cd__contains=itmtext) | Q(item_spec__contains=itmtext))[:10]

    itmstr = ""
    if rsItem:
        for i in rsItem:
            itmstr += f"<div><i class='icofont-plus-square' style='margin-right:20px;' itemid='{i.id}' bomid='{bomid}' flag='add' onclick='pickBOMItem(this)'></i>  " +\
                      f"<i class='icofont-check' style='margin-right:20px;' itemid='{i.id}' bomid='{bomid}' flag='update' onclick='pickBOMItem(this)'></i> "+\
                      f"<span>{i.item_cd} - {i.item_spec} </span></div>"

    else:
        itmstr = "<div>No item searched...</div>"

    #print(itmstr)
    context["itmstr"] = itmstr
    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bomitem_pick(request):
    context = {}

    bomid = request.GET['bomid']
    itemid = request.GET['itemid']
    flag = request.GET['flag']

    if flag == 'add':
        print(11)
        rsTmp = CMBOM.objects.get(id=bomid)
        bomorder=rsTmp.bom_order
        bomlevel=rsTmp.bom_level
        topid=rsTmp.top_id
        rsTmp.leaf_flag='0'
        rsTmp.save()

        CMBOM.objects.create(bom_type='MBOM',
                             item_id=itemid,
                             parent_id=bomid,
                             top_id=topid,
                             bom_order=bomorder+1,
                             bom_level=bomlevel+1,
                             leaf_flag='1',
                             moitem_base=0.0,
                             jaitem_base=0.0,
                             unit_product='',
                             demand_amt=0.0,
                             free_flag='1',
                             loss_product=0.0,
                             start_date='',
                             end_date='',
                             register_date=datetime.now(),
                             usage_flag='1')

        context["flag"] = "0"
        context["result_msg"] = "BOM tree added..."
        return JsonResponse(context, content_type="application/json")

    elif flag == 'update':
        rsTmp = CMBOM.objects.get(id=bomid)
        rsTmp.item_id=itemid
        rsTmp.save()

        context["flag"] = "0"
        context["result_msg"] = "BOM item updated..."
        return JsonResponse(context, content_type="application/json")
    else:
        context["flag"] = "1"
        context["result_msg"] = "Nothing..."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bom_update(request):
    context = {}

    bomid = request.GET['bomid']
    flag = request.GET['flag']
    bvalue = request.GET['bvalue']

    rsTmp = CMBOM.objects.get(id=bomid)
    if flag == 'mobase':
        rsTmp.moitem_base=bvalue
        rsTmp.save()
    elif flag == 'jabase':
        rsTmp.jaitem_base=bvalue
        rsTmp.save()
    elif flag == 'unit':
        rsTmp.unit_product=bvalue
        rsTmp.save()
    elif flag == 'loss':
        rsTmp.loss_product=bvalue
        rsTmp.save()
    elif flag == 'demand':
        rsTmp.demand_amt=bvalue
        rsTmp.save()
    elif flag == 'sdate':
        rsTmp.start_date=bvalue
        rsTmp.save()
    elif flag == 'edate':
        rsTmp.end_date=bvalue
        rsTmp.save()
    else:
        context["flag"] = "1"
        context["result_msg"] = "Nothing updated..."
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "BOM updated..."
    return JsonResponse(context, content_type="application/json")
